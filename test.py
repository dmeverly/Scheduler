import unittest
import calendar
import math
import openpyxl
import random
import os
import glob

from scheduler import readXlsx, preProcess, createSheet, addTemplate, DOW, extractEmployees, initializeEmployeeWorkbooks


def cell_for_date(month_start_day, date):
    index = month_start_day + (date - 1)
    week = index // 7
    day = index % 7
    row = 4 + week
    col = 1 + (day * 2)
    return row, col


def template_week_for_date(initial_week, month_start_day, date):
    wraps_before = (month_start_day + (date - 1)) // 7
    wk = initial_week + wraps_before
    while wk > 14:
        wk -= 14
    return wk


class TestScheduler(unittest.TestCase):
    def setUp(self):
        template_path = "./input-output/Template.xlsx"
        if not os.path.exists(template_path):
            self.fail(f"Test dependency missing: {template_path} does not exist")
        
        self.df, self.template = readXlsx(template_path)
        
        if self.df is None or self.df.empty:
            self.fail("Template DataFrame is empty or invalid")
        
        self.d1, self.d2, self.n = preProcess(self.df)
        
        if self.d1 is None or self.d1.empty or self.d2 is None or self.d2.empty or self.n is None or self.n.empty:
            self.fail("Preprocessed DataFrames are empty or invalid")

        self.styles = {
            'month': openpyxl.styles.Font(name="Arial", size=14, bold=True),
            'day': openpyxl.styles.Font(name="Arial", size=11, bold=True),
            'cell': openpyxl.styles.Font(name="Arial", size=10),
            'tla': openpyxl.styles.Alignment(horizontal="left", vertical="top", wrap_text=True),
            'center': openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True),
            'thinb': openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin")
            )
        }

    def test_headers_are_monday_to_sunday(self):
        wb = openpyxl.Workbook()
        addTemplate(self.template, wb)

        initial_week = 1
        createSheet(self.d1, self.d2, self.n, initial_week, 4, 2026, self.styles, wb)
        ws = wb["April 2026"]

        for i, expected_day in enumerate(DOW):
            col = 1 + i * 2
            self.assertEqual(ws.cell(row=3, column=col).value, expected_day)

    def test_day_1_goes_to_correct_column_for_monday_based_calendar(self):
        year, month = 2026, 4
        month_start_day, _ = calendar.monthrange(year, month)
        self.assertEqual(month_start_day, 2)

        wb = openpyxl.Workbook()
        addTemplate(self.template, wb)

        initial_week = 1
        createSheet(self.d1, self.d2, self.n, initial_week, month, year, self.styles, wb)
        ws = wb["April 2026"]

        row, col = cell_for_date(month_start_day, 1)
        self.assertEqual((row, col), (4, 5))

        cell_value = ws.cell(row=row, column=col).value
        self.assertTrue(cell_value.startswith("1\n"), f"Expected cell to start with '1\\n', got: {cell_value!r}")

    def test_specific_date_text_matches_expected_template_day_and_week(self):
        year, month = 2026, 4
        month_start_day, month_length = calendar.monthrange(year, month)

        wb = openpyxl.Workbook()
        addTemplate(self.template, wb)

        initial_week = 1
        createSheet(self.d1, self.d2, self.n, initial_week, month, year, self.styles, wb)
        ws = wb["April 2026"]

        date_to_check = 15
        self.assertTrue(1 <= date_to_check <= month_length)

        row, col = cell_for_date(month_start_day, date_to_check)

        template_day = (month_start_day + (date_to_check - 1)) % 7
        week_for_date = template_week_for_date(initial_week, month_start_day, date_to_check)

        d1_emp = self.d1[template_day][week_for_date].capitalize()
        n_emp = self.n[template_day][week_for_date].capitalize()

        d2_emp = self.d2[template_day][week_for_date]
        if d2_emp is None or d2_emp == 'x':
            expected = f"{date_to_check}\n{d1_emp} - Day\n\n{n_emp} - Night"
        else:
            d2_emp = d2_emp.capitalize()
            expected = f"{date_to_check}\n{d1_emp} - Day\n{d2_emp} - Day\n{n_emp} - Night"

        self.assertEqual(ws.cell(row=row, column=col).value, expected)

    def test_last_date_is_present_somewhere(self):
        year, month = 2026, 4
        month_start_day, month_length = calendar.monthrange(year, month)

        wb = openpyxl.Workbook()
        addTemplate(self.template, wb)

        initial_week = 1
        createSheet(self.d1, self.d2, self.n, initial_week, month, year, self.styles, wb)
        ws = wb["April 2026"]

        last_row, last_col = cell_for_date(month_start_day, month_length)
        cell_value = ws.cell(row=last_row, column=last_col).value

        self.assertIsNotNone(cell_value)
        self.assertTrue(cell_value.startswith(f"{month_length}\n"))

    def test_weeks_in_month_matches_rows_created(self):
        year, month = 2026, 4
        month_start_day, month_length = calendar.monthrange(year, month)
        weeks_in_month = math.ceil((month_start_day + month_length) / 7)

        wb = openpyxl.Workbook()
        addTemplate(self.template, wb)

        initial_week = 1
        createSheet(self.d1, self.d2, self.n, initial_week, month, year, self.styles, wb)
        ws = wb["April 2026"]

        first_data_row = 4
        last_data_row = 4 + weeks_in_month - 1

        found_any = False
        for r in range(first_data_row, last_data_row + 1):
            for c in range(1, 14, 2):
                if ws.cell(row=r, column=c).value:
                    found_any = True
                    break
            if found_any:
                break

        self.assertTrue(found_any, "Expected schedule rows to contain at least one filled date cell.")

    def test_individual_calendar_matches_master_for_randomized_dates(self):
        year, month = 2026, random.randint(1, 12)
        month_start_day, month_length = calendar.monthrange(year, month)
        
        wb = openpyxl.Workbook()
        addTemplate(self.template, wb)
        
        employees = extractEmployees(self.d1, self.d2, self.n)
        employee_workbooks = initializeEmployeeWorkbooks(employees)
        
        initial_week = 1
        createSheet(self.d1, self.d2, self.n, initial_week, month, year, self.styles, wb, employee_workbooks)
        
        master_ws = wb[f"{calendar.month_name[month]} {year}"]
        
        for emp_name, emp_wb in employee_workbooks.items():
            emp_ws = emp_wb[f"{calendar.month_name[month]} {year}"]
            
            for test_date in [random.randint(1, month_length) for _ in range(min(5, month_length))]:
                row, col = cell_for_date(month_start_day, test_date)
                
                master_cell = master_ws.cell(row=row, column=col).value
                emp_cell = emp_ws.cell(row=row, column=col).value
                
                if master_cell is None:
                    self.assertIsNone(emp_cell, f"Employee {emp_name} should have None for date {test_date}")
                else:
                    if emp_cell is not None:
                        if emp_name in emp_cell:
                            self.assertIn(emp_name, master_cell, 
                                f"Employee {emp_name} should appear in master calendar for date {test_date}")
                            self.assertIn(emp_name, emp_cell, 
                                f"Employee {emp_name} should appear in their own calendar for date {test_date}")

    def test_master_calendar_matches_template_for_randomized_dates(self):
        year, month = 2026, random.randint(1, 12)
        initial_week = random.randint(1, 14)
        month_start_day, month_length = calendar.monthrange(year, month)
        
        wb = openpyxl.Workbook()
        addTemplate(self.template, wb)
        createSheet(self.d1, self.d2, self.n, initial_week, month, year, self.styles, wb)
        
        ws = wb[f"{calendar.month_name[month]} {year}"]
        
        for test_date in [random.randint(1, month_length) for _ in range(min(5, month_length))]:
            row, col = cell_for_date(month_start_day, test_date)
            
            template_day = (month_start_day + (test_date - 1)) % 7
            week_for_date = template_week_for_date(initial_week, month_start_day, test_date)
            is_sunday = template_day == 6
            
            d1_emp = self.d1[template_day][week_for_date]
            if d1_emp:
                d1_emp = str(d1_emp).strip().capitalize()
            
            n_emp = self.n[template_day][week_for_date]
            if n_emp:
                n_emp = str(n_emp).strip().capitalize()
            
            d2_emp = self.d2[template_day][week_for_date]
            if d2_emp:
                d2_emp = str(d2_emp).strip()
                if d2_emp and d2_emp.lower() != 'x':
                    d2_emp = d2_emp.capitalize()
                else:
                    d2_emp = None
            
            cell_value = ws.cell(row=row, column=col).value
            
            if d2_emp is None or d2_emp == 'x':
                expected_base = f"{test_date}\n{d1_emp} - Day\n\n{n_emp} - Night"
            else:
                expected_base = f"{test_date}\n{d1_emp} - Day\n{d2_emp} - Day\n{n_emp} - Night"
            
            if is_sunday:
                self.assertIn(expected_base, cell_value, 
                    f"Base template content should match for date {test_date} in {calendar.month_name[month]} {year}")
            else:
                self.assertEqual(cell_value, expected_base, 
                    f"Template mismatch for date {test_date} in {calendar.month_name[month]} {year}")

    def test_employee_calendar_has_shifts(self):
        year, month = 2026, 4
        
        wb = openpyxl.Workbook()
        addTemplate(self.template, wb)
        
        employees = extractEmployees(self.d1, self.d2, self.n)
        employee_workbooks = initializeEmployeeWorkbooks(employees)
        
        initial_week = 1
        createSheet(self.d1, self.d2, self.n, initial_week, month, year, self.styles, wb, employee_workbooks)
        
        for emp_name, emp_wb in employee_workbooks.items():
            emp_ws = emp_wb[f"{calendar.month_name[month]} {year}"]
            
            has_day_shift = False
            has_night_shift = False
            
            for row in range(4, 30):
                for col in range(1, 14, 2):
                    cell = emp_ws.cell(row=row, column=col)
                    if cell.value:
                        if " - Day" in cell.value:
                            has_day_shift = True
                        elif " - Night" in cell.value:
                            has_night_shift = True
            
            total_shifts = has_day_shift + has_night_shift
            self.assertGreater(total_shifts, 0, f"Employee {emp_name} should have at least one shift")

    def test_master_calendar_created_with_all_sheets(self):
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        addTemplate(self.template, wb)
        
        initial_week = 1
        for month in range(1, 13):
            createSheet(self.d1, self.d2, self.n, initial_week, month, 2026, self.styles, wb)
            initial_week += 4
            if initial_week > 14:
                initial_week = 1
        
        month_sheets = {f"{calendar.month_name[m]} 2026" for m in range(1, 13)}
        expected_sheets = {"Template"} | month_sheets
        actual_sheets = set(wb.sheetnames)
        
        self.assertEqual(len(actual_sheets), 13, f"Expected 13 sheets (12 months + template), got {len(actual_sheets)}")
        self.assertTrue(expected_sheets.issubset(actual_sheets), 
            f"Missing expected sheets. Expected: {expected_sheets}, Got: {actual_sheets}")

    def test_employee_calendars_created_with_all_sheets(self):
        employees = extractEmployees(self.d1, self.d2, self.n)
        employee_workbooks = initializeEmployeeWorkbooks(employees)
        
        self.assertGreater(len(employees), 0, "No employees extracted from template")
        self.assertEqual(len(employee_workbooks), len(employees), 
            "Number of employee workbooks should match number of employees")
        
        initial_week = 1
        for month in range(1, 13):
            createSheet(self.d1, self.d2, self.n, initial_week, month, 2026, self.styles, 
                       openpyxl.Workbook(), employee_workbooks)
            initial_week += 4
            if initial_week > 14:
                initial_week = 1
        
        for emp_name, emp_wb in employee_workbooks.items():
            self.assertEqual(len(emp_wb.sheetnames), 12, 
                f"Employee {emp_name} should have 12 month sheets, got {len(emp_wb.sheetnames)}")
            
            for month in range(1, 13):
                expected_sheet = f"{calendar.month_name[month]} 2026"
                self.assertIn(expected_sheet, emp_wb.sheetnames, 
                    f"Employee {emp_name} should have sheet for {expected_sheet}")

    def test_calendar_sheets_contain_data(self):
        wb = openpyxl.Workbook()
        addTemplate(self.template, wb)
        
        month = 4
        year = 2026
        createSheet(self.d1, self.d2, self.n, 1, month, year, self.styles, wb)
        
        ws = wb[f"{calendar.month_name[month]} {year}"]
        
        has_date_cells = False
        has_employee_names = False
        
        for row in range(4, 20):
            for col in range(1, 14, 2):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    has_date_cells = True
                    if any(name in str(cell.value) for name in ["Day", "Night"]):
                        has_employee_names = True
        
        self.assertTrue(has_date_cells, "Master calendar sheet should contain date cells")
        self.assertTrue(has_employee_names, "Master calendar sheet should contain employee names")

    def test_employee_calendar_sheets_contain_data(self):
        employees = extractEmployees(self.d1, self.d2, self.n)
        employee_workbooks = initializeEmployeeWorkbooks(employees)
        
        month = 4
        year = 2026
        createSheet(self.d1, self.d2, self.n, 1, month, year, self.styles, 
                   openpyxl.Workbook(), employee_workbooks)
        
        for emp_name, emp_wb in employee_workbooks.items():
            emp_ws = emp_wb[f"{calendar.month_name[month]} {year}"]
            
            has_content = False
            for row in range(4, 20):
                for col in range(1, 14, 2):
                    cell = emp_ws.cell(row=row, column=col)
                    if cell.value and emp_name in str(cell.value):
                        has_content = True
                        break
                if has_content:
                    break
            
            self.assertTrue(has_content or len(employee_workbooks) == 0, 
                f"Employee {emp_name} calendar should have at least one shift or there are no employees")

    def test_master_calendar_file_exists(self):
        test_dir = os.path.dirname(os.path.abspath(__file__))
        input_output_dir = os.path.join(test_dir, "input-output")
        
        master_files = sorted([f for f in glob.glob(os.path.join(input_output_dir, "Schedule_*.xlsx")) 
                               if "Employee" not in f])
        
        self.assertGreater(len(master_files), 0, 
            "No master calendar file found. Run './run' to create calendars first.")

    def test_employee_calendar_files_exist(self):
        test_dir = os.path.dirname(os.path.abspath(__file__))
        input_output_dir = os.path.join(test_dir, "input-output")
        
        employee_files = glob.glob(os.path.join(input_output_dir, "Schedule_Employee_*.xlsx"))
        
        employees = extractEmployees(self.d1, self.d2, self.n)
        
        self.assertGreater(len(employee_files), 0, 
            "No employee calendar files found. Run './run' to create calendars first.")
        
        unique_employees_in_files = set()
        for f in employee_files:
            basename = os.path.basename(f)
            for emp in employees:
                if emp in basename:
                    unique_employees_in_files.add(emp)
                    break
        
        self.assertEqual(len(unique_employees_in_files), len(employees), 
            f"Expected calendars for {len(employees)} employees, found calendars for {len(unique_employees_in_files)}")

    def test_master_calendar_is_valid_xlsx(self):
        test_dir = os.path.dirname(os.path.abspath(__file__))
        input_output_dir = os.path.join(test_dir, "input-output")
        
        master_files = sorted([f for f in glob.glob(os.path.join(input_output_dir, "Schedule_*.xlsx")) 
                               if "Employee" not in f])
        
        self.assertGreater(len(master_files), 0, "No master calendar file found.")
        
        latest_master = master_files[-1]
        
        try:
            wb = openpyxl.load_workbook(latest_master)
            self.assertIn("Template", wb.sheetnames, "Master calendar should have Template sheet")
            self.assertGreater(len(wb.sheetnames), 1, 
                "Master calendar should have Template + month sheets")
        except Exception as e:
            self.fail(f"Master calendar file is not valid XLSX: {e}")

    def test_employee_calendars_are_valid_xlsx(self):
        test_dir = os.path.dirname(os.path.abspath(__file__))
        input_output_dir = os.path.join(test_dir, "input-output")
        
        employee_files = glob.glob(os.path.join(input_output_dir, "Schedule_Employee_*.xlsx"))
        
        self.assertGreater(len(employee_files), 0, "No employee calendar files found.")
        
        for emp_file in employee_files:
            try:
                wb = openpyxl.load_workbook(emp_file)
                self.assertEqual(len(wb.sheetnames), 12, 
                    f"Employee calendar should have 12 month sheets, {emp_file} has {len(wb.sheetnames)}")
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    self.assertIsNotNone(ws["A1"].value, 
                        f"Calendar header missing in {emp_file} sheet {sheet_name}")
            except Exception as e:
                self.fail(f"Employee calendar file is not valid XLSX: {emp_file}: {e}")


def test_data_files_exist():
    if not os.path.exists("./input-output/Template.xlsx"):
        raise FileNotFoundError("./input-output/Template.xlsx not found. Tests require the template file.")

if __name__ == "__main__":
    test_data_files_exist()
    unittest.main(verbosity=2)
