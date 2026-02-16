import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd
import calendar
import math
from datetime import datetime
import sys

TIME = datetime.now().strftime("%Y-%m-%d-%H-%M")
INPUT = "./input-output/Template.xlsx"
OUTPUT = "./input-output/Schedule_" + TIME + ".xlsx"
DOW = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
numMonths = 12


class StyleManager:
    def __init__(self):
        self.styles = {
            'month': Font(name="Arial", size=14, bold=True),
            'day': Font(name="Arial", size=11, bold=True),
            'cell': Font(name="Arial", size=10),
            'tla': Alignment(horizontal="left", vertical="top", wrap_text=True),
            'center': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'thinb': Border(bottom=Side(style="thin"), top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin")),
            'day_fill': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            'night_fill': PatternFill(start_color="ADADAD", end_color="ADADAD", fill_type="solid")
        }
    
    def get(self, key):
        return self.styles.get(key)


class ShiftCalculator:
    @staticmethod
    def build_text(date, d1_emp, d2_emp, n_emp, day, week_number, target_emp=None):
        shift_type = None
        if target_emp:
            employees_for_date = []
            if d1_emp and d1_emp == target_emp:
                employees_for_date.append(f"{d1_emp} - Day")
                shift_type = "Day"
            if d2_emp and d2_emp == target_emp:
                employees_for_date.append(f"{d2_emp} - Day")
                shift_type = "Day"
            if n_emp and n_emp == target_emp:
                employees_for_date.append(f"{n_emp} - Night")
                shift_type = "Night"
            
            if not employees_for_date:
                return "", None
            
            shift_text = f"{date}\n" + "\n".join(employees_for_date)
        else:
            if not d2_emp or d2_emp == 'x':
                shift_text = f"{date}\n{d1_emp} - Day\n\n{n_emp} - Night"
            else:
                shift_text = f"{date}\n{d1_emp} - Day\n{d2_emp} - Day\n{n_emp} - Night"
        
        if day == 6:
            shift_text += f"\nTemplate Week {week_number}"
            if target_emp is None:
                shift_text += " - Dr. Amin"
        
        return shift_text, shift_type


class ScheduleSheet:
    COLS_PER_WEEK = 14
    DATA_START_ROW = 4
    HEADER_ROW = 3
    TITLE_ROWS = 2
    COL_WIDTH = 16
    ROW_HEIGHT_BASE = 25
    ROW_HEIGHT_PER_LINE = 13
    
    def __init__(self, title, styles, month_start_day, month_length, weeks_in_month):
        self.title = title
        self.styles = styles
        self.month_start_day = month_start_day
        self.month_length = month_length
        self.weeks_in_month = weeks_in_month
    
    def apply_title(self, ws):
        ws["A1"] = self.title
        ws["A1"].font = self.styles.get('month')
        ws["A1"].alignment = self.styles.get('center')
        ws["A1"].border = self.styles.get('thinb')
        ws.merge_cells(start_row=1, start_column=1, end_row=self.TITLE_ROWS, end_column=self.COLS_PER_WEEK)
    
    def apply_headers(self, ws):
        for i, day in enumerate(DOW):
            col = i * 2 + 1
            ws.cell(row=self.HEADER_ROW, column=col).value = day
            ws.merge_cells(start_row=self.HEADER_ROW, start_column=col, end_row=self.HEADER_ROW, end_column=col + 1)
            ws.cell(row=self.HEADER_ROW, column=col).font = self.styles.get('day')
            ws.cell(row=self.HEADER_ROW, column=col).alignment = self.styles.get('center')
    
    def apply_cell_styling(self, ws, row, col, cell_value, shift_type=None):
        cell = ws.cell(row=row, column=col, value=cell_value)
        cell.alignment = self.styles.get('tla')
        cell.font = self.styles.get('cell')
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        
        for merge_col in range(col, col + 2):
            ws.cell(row=row, column=merge_col).border = self.styles.get('thinb')
            if shift_type == "Day":
                ws.cell(row=row, column=merge_col).fill = self.styles.get('day_fill')
            elif shift_type == "Night":
                ws.cell(row=row, column=merge_col).fill = self.styles.get('night_fill')
    
    def apply_borders(self, ws):
        for row in range(self.HEADER_ROW, self.DATA_START_ROW + self.weeks_in_month):
            for col in range(1, self.COLS_PER_WEEK + 1):
                ws.cell(row=row, column=col).border = self.styles.get('thinb')
    
    def set_column_widths(self, ws):
        for col in range(1, self.COLS_PER_WEEK + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = self.COL_WIDTH
    
    def set_row_heights(self, ws):
        for row in range(self.DATA_START_ROW, self.DATA_START_ROW + self.weeks_in_month):
            max_lines = max(
                ws.cell(row=row, column=col).value.count("\n") + 1 
                if ws.cell(row=row, column=col).value else 1
                for col in range(1, self.COLS_PER_WEEK, 2)
            )
            ws.row_dimensions[row].height = max(self.ROW_HEIGHT_BASE, max_lines * self.ROW_HEIGHT_PER_LINE)
    
    def set_page_setup(self, ws):
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 1
        ws.page_setup.fitToWidth = 1
    
    def format_sheet(self, ws):
        self.apply_title(ws)
        self.apply_headers(ws)
        self.apply_borders(ws)
        self.set_column_widths(ws)
        self.set_row_heights(ws)
        self.set_page_setup(ws)


def readXlsx(input_path=INPUT):
    workbook = openpyxl.load_workbook(input_path)
    workspace = workbook.active
    dataFrame = list(workspace.iter_rows(values_only=True))
    df = pd.DataFrame(data=dataFrame, dtype=str)
    return df, workspace


def buildShiftText(date, d1_emp, d2_emp, n_emp, day, weekNumber, target_emp=None):
    return ShiftCalculator.build_text(date, d1_emp, d2_emp, n_emp, day, weekNumber, target_emp)


def createSheet(d1, d2, n, weekNumber, month, year, Styles, wb, employee_workbooks=None):
    title = f"{calendar.month_name[month]} {year}"
    ws = wb.create_sheet(title=title)
    
    employee_sheets = {}
    if employee_workbooks:
        for emp_name, emp_wb in employee_workbooks.items():
            emp_ws = emp_wb.create_sheet(title=title)
            employee_sheets[emp_name] = emp_ws
    
    month_start_day, month_length = calendar.monthrange(year, month)
    month_start_day = month_start_day % 7
    weeks_in_month = math.ceil((month_start_day + month_length) / 7)
    
    sheet = ScheduleSheet(title, Styles, month_start_day, month_length, weeks_in_month)
    
    sheet.apply_title(ws)
    if employee_sheets:
        for emp_ws in employee_sheets.values():
            sheet.apply_title(emp_ws)
    
    sheet.apply_headers(ws)
    if employee_sheets:
        for emp_ws in employee_sheets.values():
            sheet.apply_headers(emp_ws)
    
    start_column = (month_start_day * 2) + 1
    date = 1
    template_day = month_start_day
    
    for week in range(weeks_in_month):
        for day in range(7):
            col = day * 2 + 1
            if week == 0 and col < start_column:
                continue
            if date > month_length:
                break
            
            d1_emp = d1[template_day][weekNumber]
            if d1_emp:
                d1_emp = str(d1_emp).strip().capitalize()
            n_emp = n[template_day][weekNumber]
            if n_emp:
                n_emp = str(n_emp).strip().capitalize()
            
            d2_emp = d2[template_day][weekNumber]
            if d2_emp:
                d2_emp = str(d2_emp).strip()
                if d2_emp and d2_emp.lower() != 'x':
                    d2_emp = d2_emp.capitalize()
                else:
                    d2_emp = None
            
            shift_text, _ = buildShiftText(date, d1_emp, d2_emp, n_emp, day, weekNumber)
            sheet.apply_cell_styling(ws, ScheduleSheet.DATA_START_ROW + week, col, shift_text)
            
            if employee_sheets:
                for emp_name, emp_ws in employee_sheets.items():
                    emp_shift_text, shift_type = buildShiftText(date, d1_emp, d2_emp, n_emp, day, weekNumber, target_emp=emp_name)
                    
                    if emp_shift_text:
                        sheet.apply_cell_styling(emp_ws, ScheduleSheet.DATA_START_ROW + week, col, emp_shift_text, shift_type)
                    else:
                        emp_ws.merge_cells(start_row=ScheduleSheet.DATA_START_ROW + week, start_column=col, 
                                          end_row=ScheduleSheet.DATA_START_ROW + week, end_column=col + 1)
                        for merge_col in range(col, col + 2):
                            emp_ws.cell(row=ScheduleSheet.DATA_START_ROW + week, column=merge_col).border = Styles.get('thinb')
            
            date += 1
            template_day += 1
            if template_day > 6:
                template_day = 0
                weekNumber += 1
                if weekNumber > 14:
                    weekNumber = 1
    
    sheet.format_sheet(ws)
    if employee_sheets:
        for emp_ws in employee_sheets.values():
            sheet.format_sheet(emp_ws)
    
    return weekNumber

def addTemplate(template, wb):
    title = "Template"
    ws = wb.create_sheet(title=title)

    for row in template.iter_rows(values_only=True):
        ws.append(row)

    for ridx, row in enumerate(
        template.iter_rows(min_row=1, max_row=template.max_row,
                           min_col=1, max_col=template.max_column),
        start=1
    ):
        for cidx, cell in enumerate(row, start=1):
            ws.cell(row=ridx, column=cidx).fill = cell.fill.copy()

def preProcess(df):
    df = df.drop([0, 1])
    df = df.iloc[:14, :]
    df = df.drop([0, 8, 16, 24], axis=1)
    df.columns = list(range(21))
    df.index = list(range(1, 15))
    d1 = df.iloc[:, :7]
    d1.columns = list(range(7))
    d2 = df.iloc[:, 7:14]
    d2.columns = list(range(7))
    n = df.iloc[:, 14:]
    n.columns = list(range(7))
    
    return d1, d2, n

def extractEmployees(d1, d2, n):
    employees = set()
    employees.update(d1.values.flatten())
    employees.update(d2.values.flatten())
    employees.update(n.values.flatten())
    employees = {str(e).strip().capitalize() for e in employees if e and str(e).strip() != 'x'}
    return sorted(employees)

def initializeEmployeeWorkbooks(employees):
    employee_workbooks = {}
    for name in employees:
        wb = Workbook()
        wb.remove(wb.active)
        employee_workbooks[name] = wb
    return employee_workbooks

def saveEmployeeWorkbooks(employee_workbooks, timestamp):
    for name, wb in employee_workbooks.items():
        safe_name = name.replace(' ', '_').replace('/', '_')
        output_path = f"./input-output/Schedule_Employee_{safe_name}_{timestamp}.xlsx"
        wb.save(output_path)
        print(f"Employee schedule saved: {output_path}")

def queryInput(string, type):
    keyboard = input(string)
    valid = False
    match type:
        case 'week':
            while(not valid):
                try:
                    value = int(keyboard)
                    if value > 0 and value < 15:
                        valid = True
                except:
                    print("Invalid Entry")
        case 'month':
            while(not valid):
                try:
                    value = int(keyboard)
                    if value > 0 and value < 13:
                        valid = True
                except:
                    print("Invalid Entry")
        case 'year':
            while(not valid):
                try:
                    value = int(keyboard)
                    if value > 2024 and value < 2030:
                        valid = True
                except:
                    print("Invalid Entry")
    return value

if __name__ == "__main__":
    df, template = readXlsx()
    d1, d2, n = preProcess(df)

    if len(sys.argv) == 4:
        weekNumber = int(sys.argv[1])
        monthStart = int(sys.argv[2])
        year = int(sys.argv[3])
        print("Using Default Input Values")
        print(f"Template Week Number: {weekNumber}")
        print(f"Starting Month Number: {monthStart}")
        print(f"Starting Year: {year}")
    else:
        weekNumber = queryInput("Template Week Number: ", "week")
        monthStart = queryInput("Starting Month Number: ", "month")
        year = queryInput("Starting Year: ", "year")

    style_manager = StyleManager()
    
    employees = extractEmployees(d1, d2, n)
    employee_workbooks = initializeEmployeeWorkbooks(employees)

    wb = Workbook()
    month = monthStart
    addTemplate(template, wb)
    for i in range(numMonths):
        if month == 13:
            month = 1
            year += 1
        weekNumber = createSheet(d1, d2, n, weekNumber, month, year, style_manager, wb, employee_workbooks)
        month += 1

    wb.remove(wb['Sheet'])
    wb.save(OUTPUT)
    print(f"Schedule saved as {OUTPUT}")
    
    saveEmployeeWorkbooks(employee_workbooks, TIME)
